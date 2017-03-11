#! /usr/bin/env python
# -*- coding: utf-8 -*-

import json
import csv
from openpyxl import load_workbook
from collections import defaultdict

class BusinessError(Exception):
    """
    Exception destinée à identifier des erreurs métier.
    """
    def __init__(self, value, cause=None):
        super(BusinessError, self)\
            .__init__(value + u', caused by ' + repr(cause))
        self.cause = cause


class Loader:

    def __init__(self, config_file):
        self.xls_file = ""
        with open(config_file) as jsonFile:
            self.config_file = json.load(jsonFile)
        self.data = defaultdict(dict)

    # --------------------------------------------#
    #       Converti la LETTRE en COLONNE NUM     #
    # --------------------------------------------#

    @staticmethod
    def lettreVersCol(lettre):
        if(len(lettre) == 1):
            return (ord(lettre.upper()) - ord('A')) + 1

    def load_file(self, xls_file):
        self.xls_file = xls_file
        wb = load_workbook(self.xls_file, data_only=True)
        ws = wb.get_sheet_by_name(self.config_file['sheet-name'])
        row = int(self.config_file['cellule-origine']['row'])
        col = int(self.config_file['cellule-origine']['col'])
        row_max = int(self.config_file['row-max'])
        # Check the column titles if they are same as the configuration
        for column in self.config_file['topology']:
            if ws[column + self.config_file['row-titles']].value != self.config_file['topology'][column]['columnTitle']:
                # Error on a column check
                raise BusinessError(
                    "column analyse of the excel file is incoherent in the cell :" + column + self.config_file['row-titles'] +
                    " find : " + ws[column + self.config_file['row-titles']].value + " instead of : " + self.config_file['topology'][column]['columnTitle'],
                    cause=self)
        # parsing file
        while row != row_max:
            for column in sorted(self.config_file['topology']):
                # Store the cell value
                cell_value = str(ws.cell(row=row, column=self.lettreVersCol(column)).value)
                if cell_value is not None:
                    # got a value
                    self.data[str(row)][self.config_file['topology'][column]['property']] = cell_value
                elif self.config_file['topology'][column]['default'] != "":
                    # No value, is there a default value for this column ?
                    self.data[str(row)][self.config_file['topology'][column]['property']] = self.config_file['topology'][column]['default']
            row += 1

    def write_csv_file(self,csv_file):
        with open(csv_file, 'wb') as csvfile:
            csv_writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            # Write header
            csv_writer.writerow([self.config_file['topology'][column]['property'] for column in sorted(self.config_file['topology'])])
            # Write all the columns
            for line in sorted(self.data.keys()):
                csv_writer.writerow([self.data[line][self.config_file['topology'][column]['property']] for column in sorted(self.config_file['topology'])])

    def get_line_numbers(self):
        return self.data.keys()

    def get_a_line(self,line_number):
        return self.data[str(line_number)]

    def __iter__(self):
        for line in sorted(self.data.keys()):
            yield self.data[line]

    def __str__(self):
        return json.dumps(self.data, indent=2)
